from __future__ import annotations

import csv
import html
import base64
import ctypes
import json
import re
import os
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import subprocess
import sys
import tempfile
import threading
import time
from time import strftime
from tkinter import filedialog, messagebox, ttk
import tkinter as tk

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo


APP_TITLE = "ICDL-Ergebnisse"
RECIPIENT = "KG_Kaufleuteteam"
CSV_COLUMNS = ["Name", "Cert-ID", "Prüfung", "Beginn", "Ende", "Ergebnis"]
DURATION_COLUMN = "Benötigte Zeit"
OUTPUT_COLUMNS = CSV_COLUMNS + [DURATION_COLUMN]
NBSP = "\u00A0"
FONT_STYLE_TOKEN = "__ICDL_FONT_STYLE__"

DESKTOP_FOLDERID = "{B4BFCC3A-DB2C-424C-B029-7FE99A87C641}"
DOWNLOADS_FOLDERID = "{374DE290-123F-4565-9164-39C4925E467B}"


def _get_app_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _get_logs_dir() -> Path:
    logs_dir = _get_app_dir() / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)
    return logs_dir


def _log_debug(message: str) -> None:
    """Schreibt einfache Laufzeitdiagnose in logs/app.log (best effort)."""
    try:
        log_file = _get_logs_dir() / "app.log"
        with log_file.open("a", encoding="utf-8") as f:
            f.write(f"[{strftime('%Y-%m-%d %H:%M:%S')}] {message}\n")
    except Exception:
        pass


@dataclass
class ExportResult:
    output_xlsx: Path
    exam_date: datetime


@dataclass
class ProcessingResult:
    rows: list[dict[str, str]]
    output_xlsx: Path
    exam_date: datetime


def _read_examinations_csv(csv_path: Path) -> list[dict[str, str]]:
    encodings = ["utf-8-sig", "cp1252", "latin-1"]
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with csv_path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f, delimiter=";")
                rows = []
                for row in reader:
                    normalized = {k.strip(): (v.strip() if isinstance(v, str) else "") for k, v in row.items() if k}
                    if any(normalized.values()):
                        rows.append(normalized)
                return rows
        except Exception as exc:  # pragma: no cover
            last_error = exc

    raise RuntimeError(f"CSV konnte nicht gelesen werden: {last_error}")


def _parse_exam_date(rows: list[dict[str, str]]) -> datetime:
    for row in rows:
        value = row.get("Beginn", "")
        if not value:
            continue
        try:
            return datetime.strptime(value, "%d.%m.%Y %H:%M")
        except ValueError:
            continue
    raise ValueError("Kein gültiges Prüfungsdatum in Spalte 'Beginn' gefunden.")


def _build_output_path(csv_path: Path) -> Path:
    mtime = datetime.fromtimestamp(csv_path.stat().st_mtime)
    file_name = f"ICDL-Ergebnisse_{mtime:%Y%m%d_%H%M%S}.xlsx"
    return csv_path.with_name(file_name)


def _find_user_excel_template() -> Path | None:
    appdata = os.getenv("APPDATA")
    candidate_paths: list[Path] = []

    if appdata:
        appdata_path = Path(appdata)
        candidate_paths.extend(
            [
                appdata_path / "Microsoft" / "Excel" / "XLSTART" / "Mappe.xltx",
                appdata_path / "Microsoft" / "Excel" / "XLSTART" / "Book.xltx",
                appdata_path / "Microsoft" / "Templates" / "Mappe.xltx",
                appdata_path / "Microsoft" / "Templates" / "Book.xltx",
            ]
        )

    user_profile = os.getenv("USERPROFILE")
    if user_profile:
        profile_path = Path(user_profile)
        candidate_paths.extend(
            [
                profile_path / "AppData" / "Roaming" / "Microsoft" / "Excel" / "XLSTART" / "Mappe.xltx",
                profile_path / "AppData" / "Roaming" / "Microsoft" / "Excel" / "XLSTART" / "Book.xltx",
            ]
        )

    for candidate in candidate_paths:
        if candidate.exists():
            return candidate

    return None


def _resolve_excel_base_font() -> Font | None:
    template_path = _find_user_excel_template()
    if template_path is None:
        return None

    try:
        wb_template = load_workbook(template_path)
        try:
            for named_style in wb_template._named_styles:
                if getattr(named_style, "name", "").lower() == "normal":
                    font_id = getattr(named_style, "fontId", None)
                    if isinstance(font_id, int) and 0 <= font_id < len(wb_template._fonts):
                        return copy(wb_template._fonts[font_id])

            if wb_template._fonts:
                return copy(wb_template._fonts[0])
            return None
        finally:
            wb_template.close()
    except Exception as exc:
        _log_debug(f"Mappe.xltx konnte nicht ausgewertet werden: {exc}")
        return None


def _parse_datetime(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%d.%m.%Y %H:%M")
    except ValueError:
        return None


def _compute_duration_minutes(row: dict[str, str]) -> int | None:
    start = _parse_datetime(row.get("Beginn", ""))
    end = _parse_datetime(row.get("Ende", ""))
    if start is None or end is None:
        return None

    minutes = int(round((end - start).total_seconds() / 60))
    if minutes < 0:
        return None
    return minutes


def _normalize_percent_text(value: str) -> str | None:
    stripped = (value or "").strip()
    if not stripped:
        return None

    match = re.fullmatch(r"([+-]?\d+(?:[.,]\d+)?)\s*%", stripped)
    if not match:
        return None

    number = match.group(1)
    return f"{number}{NBSP}%"


def _get_output_value(row: dict[str, str], column: str) -> str | int:
    if column == DURATION_COLUMN:
        minutes = _compute_duration_minutes(row)
        return "" if minutes is None else minutes

    if column == "Ergebnis":
        raw = row.get(column, "")
        normalized_percent = _normalize_percent_text(raw)
        return normalized_percent if normalized_percent is not None else raw

    return row.get(column, "")


def _is_right_aligned_value(column: str, value: str | int) -> bool:
    if column in ("Beginn", "Ende", DURATION_COLUMN):
        return True

    if isinstance(value, int):
        return True

    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return False

        if column == "Ergebnis" and _normalize_percent_text(stripped) is not None:
            return True

        if _parse_datetime(stripped) is not None:
            return True

        normalized = stripped.replace(".", "").replace(",", ".")
        try:
            float(normalized)
            return True
        except ValueError:
            return False

    return False


def _format_html_value(column: str, value: str | int) -> str:
    if column == DURATION_COLUMN and isinstance(value, int):
        return f"{value} min"
    return "" if value == "" else str(value)


def _write_excel(rows: list[dict[str, str]], output_xlsx: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ergebnisse"

    base_font = _resolve_excel_base_font()
    header_font = Font(bold=True)
    if base_font is not None:
        header_font = copy(base_font)
        header_font.bold = True

    # Kopfzeile
    ws.append(OUTPUT_COLUMNS)
    for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        # DIN 5008:2020 (hier für Tabellenkopfzeile umgesetzt):
        # - Erste Spaltenüberschrift: linksbündig + vertikal zentriert
        # - Alle weiteren Spaltenüberschriften: horizontal + vertikal zentriert
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Daten
    for row in rows:
        ws.append([_get_output_value(row, col) for col in OUTPUT_COLUMNS])

    # Ausrichtung: Zahlen/Datum rechts, Texte links
    for row_idx in range(2, ws.max_row + 1):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if base_font is not None:
                cell.font = copy(base_font)

            if _is_right_aligned_value(col_name, cell.value):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

            if col_name == DURATION_COLUMN and isinstance(cell.value, int):
                # Benutzerdefiniertes Zahlenformat: Ganzzahl + " min"
                cell.number_format = '0 "min"'

    # Tabelle für bessere Weiterverarbeitung in Excel
    end_row = ws.max_row
    end_col = ws.max_column
    if end_row >= 2:
        table_ref = f"A1:{chr(64 + end_col)}{end_row}"
        tab = Table(displayName="ICDLErgebnisse", ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    # Spaltenbreiten
    widths = {
        "A": 28,
        "B": 14,
        "C": 28,
        "D": 18,
        "E": 18,
        "F": 10,
        "G": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_xlsx)


def _create_html_table(rows: list[dict[str, str]]) -> str:
    header_cells_parts = []
    for idx, col in enumerate(OUTPUT_COLUMNS):
        if idx == 0:
            header_cells_parts.append(
                (
                    "<th style=\"text-align:left;vertical-align:middle;"
                    f"border:1px solid #000;padding:4px;{FONT_STYLE_TOKEN}\">"
                    f"{html.escape(col)}</th>"
                )
            )
        else:
            header_cells_parts.append(
                (
                    "<th style=\"text-align:center;vertical-align:middle;"
                    f"border:1px solid #000;padding:4px;{FONT_STYLE_TOKEN}\">"
                    f"{html.escape(col)}</th>"
                )
            )
    header_cells = "".join(header_cells_parts)

    body_rows = []

    for row in rows:
        td_parts = []
        for col in OUTPUT_COLUMNS:
            value = _get_output_value(row, col)
            text = html.escape(_format_html_value(col, value))
            if _is_right_aligned_value(col, value):
                td_parts.append(
                    (
                        "<td style=\"text-align:right;vertical-align:middle;"
                        f"border:1px solid #000;padding:4px;{FONT_STYLE_TOKEN}\">"
                        f"{text}</td>"
                    )
                )
            else:
                td_parts.append(
                    (
                        "<td style=\"text-align:left;vertical-align:middle;"
                        f"border:1px solid #000;padding:4px;{FONT_STYLE_TOKEN}\">"
                        f"{text}</td>"
                    )
                )

        tds = "".join(td_parts)
        body_rows.append(f"<tr>{tds}</tr>")

    body = "".join(body_rows)
    return (
        "<table border='1' cellspacing='0' cellpadding='4' "
        f"style=\"border-collapse:collapse;border:1px solid #000;{FONT_STYLE_TOKEN}\">"
        f"<thead style=\"background:#EAF2FF;\"><tr>{header_cells}</tr></thead>"
        f"<tbody>{body}</tbody></table>"
    )


def _create_outlook_preview_mail(rows: list[dict[str, str]], exam_date: datetime, output_xlsx: Path) -> None:
    subject_date = exam_date.strftime("%d.%m.%Y")
    table_html = _create_html_table(rows)

    full_html = (
        "<p>Hallo zusammen,</p>"
        "<p>anbei die Ergebnisse der letzten ICDL-Prüfung.</p>"
        f"{table_html}"
        "<p>Viele Grüße</p>"
    )

    subject = f"ICDL-Ergebnisse der Prüfung vom {subject_date}"
    app_log = _get_logs_dir() / "app.log"
    temp_dir = Path(tempfile.gettempdir())
    stamp = f"{datetime.now():%Y%m%d_%H%M%S_%f}"
    status_file = temp_dir / f"icdl_outlook_status_{stamp}.txt"
    payload_file = temp_dir / f"icdl_outlook_payload_{stamp}.json"
    ps_file = temp_dir / f"icdl_outlook_{stamp}.ps1"

    payload = {
        "to": RECIPIENT,
        "subject": subject,
        "html": full_html,
        "attachment": str(output_xlsx),
        "status": str(status_file),
        "log": str(app_log),
    }
    payload_file.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    ps_script = """
param([Parameter(Mandatory=$true)][string]$PayloadPath)
$ErrorActionPreference = 'Stop'
try {
    $payload = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
    $outlook = $null
    try { $outlook = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Outlook.Application') } catch {}
    if ($null -eq $outlook) {
        # Outlook ist noch nicht gestartet -> automatisch starten.
        try { $outlook = New-Object -ComObject Outlook.Application } catch {}
    }

    if ($null -eq $outlook) {
        # Letzter Fallback: Outlook-Prozess starten und erneut anbinden.
        try { Start-Process -FilePath 'outlook.exe' | Out-Null } catch {}
        $deadlineOutlook = (Get-Date).AddSeconds(25)
        while ($null -eq $outlook -and (Get-Date) -lt $deadlineOutlook) {
            Start-Sleep -Milliseconds 500
            try { $outlook = [System.Runtime.InteropServices.Marshal]::GetActiveObject('Outlook.Application') } catch {}
        }
    }

    if ($null -eq $outlook) {
        throw 'Outlook konnte nicht automatisch gestartet werden. Bitte Outlook einmal manuell öffnen und erneut versuchen.'
    }

    # MAPI-Session sicherstellen (ohne erzwungene Dialoge)
    try {
        $namespace = $outlook.GetNamespace('MAPI')
        if ($null -ne $namespace) {
            try { [void]$namespace.Logon($null, $null, $false, $false) } catch {}
        }
    } catch {}

    # Outlook kann unmittelbar nach Start kurz beschäftigt sein (RPC_E_CALL_REJECTED).
    $mail = $null
    $deadlineMail = (Get-Date).AddSeconds(20)
    while ($null -eq $mail -and (Get-Date) -lt $deadlineMail) {
        try {
            $mail = $outlook.CreateItem(0)
        }
        catch {
            Start-Sleep -Milliseconds 400
        }
    }

    if ($null -eq $mail) {
        throw 'Outlook ist gestartet, aber derzeit beschäftigt (z. B. Dialog/Profil-Initialisierung). Bitte Outlook kurz in den Vordergrund bringen und erneut versuchen.'
    }

    $mail.To = $payload.to
    $mail.Subject = $payload.subject
    if (-not [string]::IsNullOrWhiteSpace($payload.attachment) -and (Test-Path -LiteralPath $payload.attachment)) {
        [void]$mail.Attachments.Add($payload.attachment)
    }

    if ([string]::IsNullOrWhiteSpace($payload.attachment) -or -not (Test-Path -LiteralPath $payload.attachment)) {
        throw 'Die erzeugte Excel-Datei für die E-Mail-Kopie wurde nicht gefunden.'
    }

    $excel = $null
    $workbook = $null
    $worksheet = $null
    $usedRange = $null
    try {
        # Excel-Quelle öffnen und für Paste mit Quellformatierung in die Zwischenablage kopieren.
        $excel = New-Object -ComObject Excel.Application
        $excel.Visible = $false
        $excel.DisplayAlerts = $false
        $workbook = $excel.Workbooks.Open($payload.attachment)
        $worksheet = $workbook.Worksheets.Item(1)
        $usedRange = $worksheet.UsedRange
        if ($null -eq $usedRange) {
            throw 'Die Excel-Tabelle enthält keinen kopierbaren Bereich.'
        }
        [void]$usedRange.Copy()
    }
    catch {
        throw ('Excel-Kopie konnte nicht vorbereitet werden: ' + $_.Exception.Message)
    }

    $mail.Display()
    $inspector = $mail.GetInspector()
    if ($null -eq $inspector) { throw 'Outlook-Vorschau konnte nicht erzeugt werden (Inspector ist null).' }

    $wordEditor = $inspector.WordEditor
    if ($null -eq $wordEditor -or $null -eq $wordEditor.Application -or $null -eq $wordEditor.Application.Selection) {
        throw 'Outlook-Editor konnte nicht initialisiert werden.'
    }

    $selection = $wordEditor.Application.Selection

    # Compose-Standardfont aus Word-Editor bestimmen (Normal/Standard-Stil bevorzugt).
    $composeStyle = $null
    $composeFontName = $null
    $composeFontSize = $null
    try {
        try { $composeStyle = $wordEditor.Styles.Item('Normal') } catch {}
        if ($null -eq $composeStyle) {
            try { $composeStyle = $wordEditor.Styles.Item('Standard') } catch {}
        }

        if ($null -ne $composeStyle -and $null -ne $composeStyle.Font) {
            $composeFontName = $composeStyle.Font.Name
            $composeFontSize = $composeStyle.Font.Size
        }
    } catch {}

    if ([string]::IsNullOrWhiteSpace($composeFontName)) {
        try { $composeFontName = $selection.Font.Name } catch {}
    }
    if ($null -eq $composeFontSize -or [double]$composeFontSize -le 0) {
        try { $composeFontSize = $selection.Font.Size } catch {}
    }
    if ([string]::IsNullOrWhiteSpace($composeFontName)) { $composeFontName = 'Calibri' }
    if ($null -eq $composeFontSize -or [double]$composeFontSize -le 0) { $composeFontSize = 11 }

    $applyComposeTextStyle = {
        param($sel, $styleObj, $fontName, $fontSize)
        try {
            if ($null -ne $styleObj) { $sel.Style = $styleObj }
        } catch {}
        try { $sel.Font.Name = $fontName } catch {}
        try { $sel.Font.Size = $fontSize } catch {}
    }

    # Inhalt an den Anfang setzen; bestehende Signatur bleibt darunter erhalten.
    $selection.SetRange(0, 0)
    & $applyComposeTextStyle $selection $composeStyle $composeFontName $composeFontSize
    $selection.TypeText('Hallo zusammen,')
    $selection.TypeParagraph()
    $selection.TypeParagraph()
    $selection.TypeText('anbei die Ergebnisse der letzten ICDL-Prüfung.')
    $selection.TypeParagraph()
    $selection.TypeParagraph()

    try {
        # Entspricht: Einfügen -> Ursprüngliche Formatierung beibehalten.
        # Parameter: LinkedToExcel=$false, WordFormatting=$false, RTF=$false
        [void]$selection.PasteExcelTable($false, $false, $false)
    }
    catch {
        # Fallback, falls PasteExcelTable in der Umgebung nicht verfügbar ist.
        [void]$selection.Paste()
    }

    $selection.TypeParagraph()
    $selection.TypeParagraph()
    & $applyComposeTextStyle $selection $composeStyle $composeFontName $composeFontSize
    $selection.TypeText('Viele Grüße')

    # Excel-Objekte sauber schließen/freigeben.
    try { if ($workbook -ne $null) { [void]$workbook.Close($false) } } catch {}
    try { if ($excel -ne $null) { [void]$excel.Quit() } } catch {}
    try { if ($usedRange -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($usedRange) } } catch {}
    try { if ($worksheet -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) } } catch {}
    try { if ($workbook -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) } } catch {}
    try { if ($excel -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) } } catch {}

    Set-Content -Path $payload.status -Value 'OK' -Encoding UTF8
    try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($inspector) } catch {}
    try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($mail) } catch {}
    try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($outlook) } catch {}
    exit 0
}
catch {
    try { Set-Content -Path $payload.status -Value ('ERR: ' + $_.Exception.Message) -Encoding UTF8 } catch {}
    try { Add-Content -Path $payload.log -Value ('[' + (Get-Date -Format 'yyyy-MM-dd HH:mm:ss') + '] Outlook-PS-Fehler: ' + $_.Exception.Message) -Encoding UTF8 } catch {}
    exit 1
}
""".strip()
    # Windows PowerShell 5.1 liest Skripte mit UTF-8-BOM zuverlässiger.
    # So bleiben Umlaute in Fehlermeldungen korrekt.
    ps_file.write_text(ps_script, encoding="utf-8-sig")

    try:
        process = subprocess.Popen(
            [
                "powershell",
                "-NoProfile",
                "-STA",
                "-ExecutionPolicy",
                "Bypass",
                "-WindowStyle",
                "Hidden",
                "-File",
                str(ps_file),
                "-PayloadPath",
                str(payload_file),
            ],
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )

        deadline = time.time() + 90
        while time.time() < deadline:
            if status_file.exists():
                # PowerShell kann je nach Version/Befehl UTF-8 mit BOM schreiben.
                # utf-8-sig entfernt ein ggf. vorhandenes BOM, damit "OK" zuverlässig erkannt wird.
                status_text = status_file.read_text(encoding="utf-8-sig", errors="ignore").strip()
                if status_text == "OK":
                    return
                raise RuntimeError(f"Outlook konnte keine Vorschau erzeugen: {status_text}")

            if process.poll() is not None:
                if process.returncode == 0:
                    _log_debug("Outlook-Vorschau bestätigt (PowerShell ExitCode 0 ohne Statusdatei).")
                    return
                raise RuntimeError("Outlook konnte keine Vorschau erzeugen (PowerShell beendet mit Fehlercode).")

            time.sleep(0.2)

        raise RuntimeError(
            "Outlook antwortet nicht rechtzeitig. Bitte Outlook einmal manuell öffnen und erneut versuchen."
        )
    finally:
        for temp_file in (status_file, payload_file, ps_file):
            try:
                temp_file.unlink(missing_ok=True)
            except Exception:
                pass


def process_file(csv_path: Path) -> ExportResult:
    result = process_csv_to_excel(csv_path)
    _create_outlook_preview_mail(result.rows, result.exam_date, result.output_xlsx)
    archived_xlsx = _archive_output_xlsx(result.output_xlsx)
    return ExportResult(output_xlsx=archived_xlsx, exam_date=result.exam_date)


def process_csv_to_excel(csv_path: Path) -> ProcessingResult:
    rows = _read_examinations_csv(csv_path)
    if not rows:
        raise ValueError("Die CSV-Datei enthält keine Datenzeilen.")

    exam_date = _parse_exam_date(rows)
    output_xlsx = _build_output_path(csv_path)
    _write_excel(rows, output_xlsx)

    return ProcessingResult(rows=rows, output_xlsx=output_xlsx, exam_date=exam_date)


def _archive_output_xlsx(output_xlsx: Path) -> Path:
    """Verschiebt die erzeugte Excel-Datei nach ./archive und überschreibt bei Bedarf."""
    archive_dir = output_xlsx.parent / "archive"
    archive_dir.mkdir(parents=True, exist_ok=True)
    archive_target = archive_dir / output_xlsx.name

    try:
        if output_xlsx.resolve() == archive_target.resolve():
            return output_xlsx
    except Exception:
        pass

    os.replace(output_xlsx, archive_target)
    return archive_target


def _get_windows_known_folder_path(folder_id: str) -> Path | None:
    """Liest bekannte Windows-Ordner (z. B. Desktop/Downloads) robust aus."""
    if not sys.platform.startswith("win"):
        return None

    class GUID(ctypes.Structure):
        _fields_ = [
            ("Data1", ctypes.c_uint32),
            ("Data2", ctypes.c_uint16),
            ("Data3", ctypes.c_uint16),
            ("Data4", ctypes.c_ubyte * 8),
        ]

    try:
        import uuid

        parsed = uuid.UUID(folder_id)
        guid = GUID()
        guid.Data1 = parsed.fields[0]
        guid.Data2 = parsed.fields[1]
        guid.Data3 = parsed.fields[2]
        tail = parsed.bytes[8:]
        for idx, value in enumerate(tail):
            guid.Data4[idx] = value

        path_ptr = ctypes.c_wchar_p()
        shell32 = ctypes.windll.shell32
        ole32 = ctypes.windll.ole32

        hresult = shell32.SHGetKnownFolderPath(ctypes.byref(guid), 0, None, ctypes.byref(path_ptr))
        if hresult != 0 or not path_ptr.value:
            return None

        folder_path = Path(path_ptr.value)
        ole32.CoTaskMemFree(path_ptr)
        return folder_path
    except Exception:
        return None


def _get_windows_desktop_downloads_dirs() -> list[Path]:
    """Ermittelt Desktop/Downloads inkl. OneDrive-Umleitungen."""
    home_dir = Path.home()
    candidates: list[Path] = []

    desktop_known = _get_windows_known_folder_path(DESKTOP_FOLDERID)
    downloads_known = _get_windows_known_folder_path(DOWNLOADS_FOLDERID)
    if desktop_known is not None:
        candidates.append(desktop_known)
    if downloads_known is not None:
        candidates.append(downloads_known)

    # Fallbacks (inkl. häufiger OneDrive-Varianten)
    candidates.extend([home_dir / "Desktop", home_dir / "Downloads"])
    for env_var in ("OneDrive", "OneDriveCommercial", "OneDriveConsumer"):
        one_drive_root = os.getenv(env_var)
        if one_drive_root:
            root_path = Path(one_drive_root)
            candidates.append(root_path / "Desktop")
            candidates.append(root_path / "Downloads")

    return candidates


def _get_csv_search_dirs() -> list[Path]:
    """Reihenfolge der CSV-Suche: App-/EXE-Ordner, Desktop, Downloads."""
    app_dir = _get_app_dir()
    candidates = [app_dir]
    if sys.platform.startswith("win"):
        candidates.extend(_get_windows_desktop_downloads_dirs())
    else:
        home_dir = Path.home()
        candidates.extend([home_dir / "Desktop", home_dir / "Downloads"])

    dirs: list[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        normalized = str(candidate.resolve()) if candidate.exists() else str(candidate)
        if normalized in seen:
            continue
        seen.add(normalized)
        dirs.append(candidate)
    return dirs


def _iter_examinations_csv_candidates() -> list[Path]:
    return [search_dir / "examinations.csv" for search_dir in _get_csv_search_dirs()]


def _resolve_initial_csv_dir() -> Path:
    for candidate in _iter_examinations_csv_candidates():
        if candidate.exists():
            return candidate.parent
    return _get_app_dir()


def _find_daily_csv_candidate() -> Path | None:
    for candidate in _iter_examinations_csv_candidates():
        if _is_file_modified_today(candidate):
            return candidate
    return None


def _resolve_docs_dir() -> Path | None:
    """Ermittelt den docs-Ordner für Source- oder EXE-Betrieb."""
    app_dir = _get_app_dir()
    candidates = [
        app_dir / "docs",
        Path(__file__).resolve().parent / "docs",
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_dir():
            return candidate
    return None


def _resolve_window_icon_path() -> Path | None:
    """Ermittelt eine vorhandene ICO-Datei für das Fenster-Icon."""
    runtime_dir = Path(getattr(sys, "_MEIPASS", _get_app_dir()))
    app_dir = _get_app_dir()
    candidates = [
        runtime_dir / "app_icon.ico",
        app_dir / "app_icon.ico",
        Path(__file__).resolve().parent / "app_icon.ico",
    ]

    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def _apply_window_icon(root: tk.Tk) -> None:
    """Setzt das Icon der Titelleiste (best effort, primär Windows)."""
    icon_path = _resolve_window_icon_path()

    if icon_path is not None:
        try:
            root.iconbitmap(default=str(icon_path))
            _log_debug(f"Fenster-Icon gesetzt: {icon_path}")
            return
        except Exception as exc:
            _log_debug(f"Fenster-Icon aus app_icon.ico konnte nicht gesetzt werden: {exc}")

    # Fallback: Im EXE-Betrieb das EXE-Icon nutzen.
    if getattr(sys, "frozen", False):
        try:
            root.iconbitmap(default=str(Path(sys.executable)))
            _log_debug("Fenster-Icon aus EXE verwendet.")
            return
        except Exception as exc:
            _log_debug(f"Fenster-Icon aus EXE konnte nicht gesetzt werden: {exc}")


def _pick_csv_file(initial_dir: Path) -> str:
    """Öffnet den CSV-Dialog mit sinnvoller Vorbelegung.

        - Sucht `examinations.csv` in dieser Reihenfolge:
            1) App-/EXE-Verzeichnis
            2) Benutzerspezifischer Desktop
            3) Benutzerspezifischer Downloads-Ordner
        - Wenn in `initial_dir` eine `examinations.csv` vorhanden ist, wird diese vorbelegt.
    - Andernfalls wird nur in das Verzeichnis geführt und die Datei frei durchsucht.
    """
    default_csv = initial_dir / "examinations.csv"

    dialog_kwargs = {
        "title": "examinations.csv auswählen",
        "initialdir": str(initial_dir),
        "filetypes": [("CSV-Dateien", "*.csv"), ("Alle Dateien", "*.*")],
    }

    if default_csv.exists():
        dialog_kwargs["initialfile"] = default_csv.name

    return filedialog.askopenfilename(**dialog_kwargs)


def _is_file_modified_today(path: Path) -> bool:
    if not path.exists():
        return False
    modified = datetime.fromtimestamp(path.stat().st_mtime)
    return modified.date() == datetime.now().date()


def run_gui() -> None:
    root = tk.Tk()
    root.title(APP_TITLE)
    _apply_window_icon(root)
    root.geometry("700x500")
    root.resizable(False, False)
    root.configure(bg="#F3F6FB")

    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure("App.TFrame", background="#F3F6FB")
    style.configure("Card.TFrame", background="#FFFFFF", relief="solid", borderwidth=1)
    style.configure("Header.TFrame", background="#EEF4FF")
    style.configure("Action.TFrame", background="#FFFFFF")
    style.configure("Title.TLabel", background="#EEF4FF", foreground="#111827", font=("Segoe UI", 15, "bold"))
    style.configure("Subtitle.TLabel", background="#EEF4FF", foreground="#4B5563", font=("Segoe UI", 10))
    style.configure("Body.TLabel", background="#FFFFFF", foreground="#374151", font=("Segoe UI", 10))
    style.configure("Section.TLabel", background="#FFFFFF", foreground="#1F2937", font=("Segoe UI", 10, "bold"))
    style.configure("Badge.TLabel", background="#DCE8FF", foreground="#1D4ED8", font=("Segoe UI", 9, "bold"), padding=(8, 4))
    style.configure("Status.TLabel", background="#FFFFFF", foreground="#111827", font=("Segoe UI", 10, "bold"))
    style.configure(
        "Primary.TButton",
        font=("Segoe UI", 10, "bold"),
        padding=(12, 9),
    )
    style.map("Primary.TButton", background=[("active", "#2563EB")])
    style.configure("Secondary.TButton", font=("Segoe UI", 10), padding=(12, 8))
    style.configure("App.TLabelframe", background="#FFFFFF", borderwidth=0)
    style.configure("App.TLabelframe.Label", background="#FFFFFF", foreground="#1F2937", font=("Segoe UI", 10, "bold"))

    frame = ttk.Frame(root, style="App.TFrame", padding=18)
    frame.pack(fill="both", expand=True)

    card = ttk.Frame(frame, style="Card.TFrame", padding=0)
    card.pack(fill="both", expand=True)

    header = ttk.Frame(card, style="Header.TFrame", padding=(16, 14))
    header.pack(fill="x")

    title_row = ttk.Frame(header, style="Header.TFrame")
    title_row.pack(fill="x")

    title_label = ttk.Label(title_row, text=APP_TITLE, style="Title.TLabel", anchor="w")
    title_label.pack(side="left", fill="x", expand=True)

    badge = ttk.Label(title_row, text="AUTOMATISIERT", style="Badge.TLabel")
    badge.pack(side="right", anchor="e")

    subtitle_label = ttk.Label(
        header,
        text="CSV einlesen, Excel erzeugen und Outlook-Vorschau automatisch vorbereiten",
        style="Subtitle.TLabel",
        anchor="w",
    )
    subtitle_label.pack(fill="x", pady=(4, 0))

    content = ttk.Frame(card, style="Card.TFrame", padding=(16, 14))
    content.pack(fill="both", expand=True)

    info_header = ttk.Label(content, text="Ablauf", style="Section.TLabel", anchor="w")
    info_header.pack(fill="x", pady=(0, 6))

    info_box = ttk.LabelFrame(content, style="App.TLabelframe", text="Schritte", padding=(12, 10))
    info_box.pack(fill="x", pady=(0, 12))

    info = ttk.Label(
        info_box,
        text=(
            "1) Tagesaktuelle examinations.csv wird automatisch verarbeitet\n"
            "   (App-Ordner, Desktop inkl. OneDrive, Downloads)\n"
            "2) Alternativ: CSV manuell auswählen\n"
            "3) Excel-Datei erzeugen\n"
            "4) Outlook-Mail mit Tabellenkopie in Vorschau öffnen"
        ),
        style="Body.TLabel",
        wraplength=620,
        justify="left",
        anchor="w",
    )
    info.pack(fill="x")

    style.configure(
        "Processing.Horizontal.TProgressbar",
        troughcolor="#E5E7EB",
        background="#2D6CDF",
    )
    style.configure(
        "Success.Horizontal.TProgressbar",
        troughcolor="#E5E7EB",
        background="#1F9D55",
    )
    style.configure(
        "Error.Horizontal.TProgressbar",
        troughcolor="#E5E7EB",
        background="#C62828",
    )

    status_var = tk.StringVar(value="Bereit")
    running_flag = {"busy": False, "phase": "idle"}
    progress_var = tk.DoubleVar(value=0.0)

    def _set_progress(value: float, style_name: str = "Processing.Horizontal.TProgressbar") -> None:
        progress.configure(style=style_name)
        progress_var.set(max(0.0, min(100.0, value)))

    def _set_busy(phase: str, status_text: str) -> None:
        running_flag["busy"] = True
        running_flag["phase"] = phase
        run_btn.state(["disabled"])
        docs_btn.state(["disabled"])
        if phase == "csv":
            _set_progress(35.0)
        elif phase == "outlook":
            _set_progress(80.0)
        else:
            _set_progress(progress_var.get())
        status_var.set(status_text)

    def _set_idle(status_text: str) -> None:
        running_flag["busy"] = False
        running_flag["phase"] = "idle"
        run_btn.state(["!disabled"])
        docs_btn.state(["!disabled"])
        if status_text.startswith("Fertig"):
            _set_progress(100.0, "Success.Horizontal.TProgressbar")
        elif status_text == "Fehler":
            _set_progress(100.0, "Error.Horizontal.TProgressbar")
        else:
            _set_progress(0.0)
        status_var.set(status_text)

    def _schedule_outlook_watchdog() -> None:
        def _watchdog() -> None:
            if running_flag.get("busy") and running_flag.get("phase") == "outlook":
                status_var.set("Outlook wird gestartet ... (bitte kurz warten)")
        root.after(6000, _watchdog)

    def _start_processing(csv_file: str, *, auto_triggered: bool = False) -> None:
        if running_flag["busy"]:
            return

        csv_path = Path(csv_file)
        if auto_triggered:
            _set_busy("csv", "Tagesaktuelle CSV erkannt – starte automatische Verarbeitung ...")
        else:
            _set_busy("csv", "Verarbeite CSV ...")

        _log_debug(f"Start Verarbeitung für Datei: {csv_file}")

        def csv_worker() -> None:
            try:
                result = process_csv_to_excel(csv_path)
                _log_debug(f"CSV/Excel abgeschlossen: {result.output_xlsx}")

                def start_outlook_phase() -> None:
                    _set_busy("outlook", "Excel erstellt – öffne Outlook-Vorschau ...")
                    _schedule_outlook_watchdog()

                    def outlook_worker() -> None:
                        try:
                            _create_outlook_preview_mail(result.rows, result.exam_date, result.output_xlsx)
                            archived_xlsx = _archive_output_xlsx(result.output_xlsx)
                            _log_debug(f"Outlook-Vorschau erfolgreich geöffnet; Excel archiviert: {archived_xlsx}")

                            def on_ok() -> None:
                                _set_idle(f"Fertig: archive\\{archived_xlsx.name}")

                            root.after(0, on_ok)
                        except Exception as exc:
                            _log_debug(f"Outlook-Fehler: {exc}")
                            err = str(exc)

                            def on_err() -> None:
                                _set_idle("Fehler")
                                messagebox.showerror(APP_TITLE, f"Fehler bei Outlook-Vorschau: {err}")

                            root.after(0, on_err)

                    threading.Thread(target=outlook_worker, daemon=True).start()

                root.after(0, start_outlook_phase)
            except Exception as exc:
                error_message = str(exc)
                _log_debug(f"CSV/Excel-Fehler: {error_message}")

                def on_error() -> None:
                    _set_idle("Fehler")
                    messagebox.showerror(APP_TITLE, f"Fehler: {error_message}")

                root.after(0, on_error)

        threading.Thread(target=csv_worker, daemon=True).start()

    def on_run() -> None:
        if running_flag["busy"]:
            return

        initial_dir = _resolve_initial_csv_dir()

        csv_file = _pick_csv_file(initial_dir)
        if not csv_file:
            return
        _start_processing(csv_file)

    def on_open_docs() -> None:
        docs_dir = _resolve_docs_dir()
        if docs_dir is None:
            messagebox.showerror(APP_TITLE, "Der Ordner 'docs' wurde nicht gefunden.")
            _log_debug("docs-Ordner konnte nicht gefunden werden.")
            return

        try:
            if sys.platform.startswith("win") and hasattr(os, "startfile"):
                os.startfile(str(docs_dir))
            else:
                subprocess.Popen(["explorer", str(docs_dir)])
            _log_debug(f"docs-Ordner geöffnet: {docs_dir}")
        except Exception as exc:
            _log_debug(f"Öffnen des docs-Ordners fehlgeschlagen: {exc}")
            messagebox.showerror(APP_TITLE, f"Der docs-Ordner konnte nicht geöffnet werden:\n{exc}")

    action_frame = ttk.Frame(content, style="Action.TFrame")
    action_frame.pack(fill="x", pady=(4, 10))

    run_btn = ttk.Button(
        action_frame,
        text="CSV einlesen und Mail-Vorschau erzeugen",
        command=on_run,
        style="Primary.TButton",
    )
    run_btn.pack(side="left", fill="x", expand=True, padx=(0, 6))

    docs_btn = ttk.Button(
        action_frame,
        text="Dokumentation öffnen",
        command=on_open_docs,
        style="Secondary.TButton",
    )
    docs_btn.pack(side="left", fill="x", expand=True, padx=(6, 0))

    status = ttk.Label(content, textvariable=status_var, style="Status.TLabel", anchor="w")
    status.pack(fill="x", pady=(2, 0))

    progress = ttk.Progressbar(
        content,
        orient="horizontal",
        mode="determinate",
        maximum=100,
        variable=progress_var,
        style="Processing.Horizontal.TProgressbar",
    )
    progress.pack(fill="x", pady=(8, 2))

    # Auto-Start: Tagesaktuelle examinations.csv in App-/EXE-Verzeichnis, Desktop oder Downloads verarbeiten.
    def _auto_start_if_daily_csv_available() -> None:
        daily_csv = _find_daily_csv_candidate()
        if daily_csv is not None:
            _log_debug(f"Auto-Start mit tagesaktueller CSV: {daily_csv}")
            _start_processing(str(daily_csv), auto_triggered=True)

    root.after(200, _auto_start_if_daily_csv_available)

    root.mainloop()


if __name__ == "__main__":
    run_gui()
